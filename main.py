from tkinter import *
from tkinter import messagebox
import openpyxl
from matplotlib.ticker import FormatStrFormatter
import xlwings
import tkinter as tk
from tkinter import filedialog
import matplotlib.pyplot as plt
from reportlab.lib import styles
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Image, Paragraph, Spacer
from PIL import Image as PILImage, ImageTk
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
file_name = "graph.png'"

if os.path.isfile(file_name):
    os.remove(file_name)
else:
    print("")


filename = "SCL-90.xlsx"
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

# Очищаем значения в ячейках с C2 по G91
for row in sheet.iter_rows(min_row=2, min_col=3, max_row=91, max_col=7):
    for cell in row:
        cell.value = None

# Сохраняем изменения
workbook.save(filename)
class Test:
    def __init__(self, master):
        self.parent = master


        self.questions = [
            "1. Головные боли",
            "2. Нервозность или внутренняя дрожь",
            "3. Повторяющиеся неприятные неотвязные мысли",
            "4. Слабость или головокружение",
            "5. Потеря сексуального влечения или удовольствия",
            "6. Чувство недовольства другими",
            "7. Ощущение, что кто-то другой может управлять Вашими мыслями",
            "8. Ощущение, что почти во всех Ваших неприятностях виноваты другие",
            "9. Проблемы с памятью",
            "10. Ваша небрежность или неряшливость",
            "11. Легко возникающая досада или раздражение",
            "12. Боли в сердце или грудной клетке",
            "13. Чувство страха в открытых местах или на улице",
            "14. Упадок сил или заторможенность",
            "15. Мысли о том, чтобы покончить с собой",
            "16. То, что Вы слышите голоса, которых не слышат другие",
            "17. Дрожь",
            "18. Чувство, что большинству людей нельзя доверять",
            "19. Плохой аппетит",
            "20. Слезливость",
            "21. Застенчивость или скованность в общении с лицами другого пола",
            "22. Ощущение, что Вы в западне или пойманы",
            "23. Неожиданный и беспричинный страх",
            "24. Вспышки гнева, которые Вы не могли сдержать",
            "25. Боязнь выйти из дома одному",
            "26. Чувство, что Вы сами во всем виноваты",
            "27. Боли в пояснице",
            "28. Ощущение, что что-то мешает Вам сделать что-либо",
            "29. Чувство одиночества",
            "30. Подавленное настроение, «хандра»",
            "31. Чрезмерное беспокойство по разным поводам",
            "32. Отсутствие интереса к чему бы то ни было",
            "33. Чувство страха",
            "34. То, что Ваши чувства легко задеть",
            "35. Ощущение, что другие проникают в Ваши мысли",
            "36. Ощущение, что другие не понимают Вас или не сочувствуют Вам",
            "37. Ощущение, что люди недружелюбны или вы им не нравитесь",
            "38. Необходимость делать все очень медленно, чтобы не допустить ошибки",
            "39. Сильное или учащенное сердцебиение",
            "40. Тошнота или расстройство желудка",
            "41. Ощущение, что Вы хуже других",
            "42. Боли в мышцах",
            "43. Ощущение, что другие наблюдают за Вами или говорят о Вас",
            "44. То, что Вам трудно заснуть",
            "45. Потребность проверять и перепроверять то, что Вы делаете",
            "46. Трудности в принятии решения",
            "47. Боязнь езды в автобусах, метро или поездах",
            "48. Затрудненное дыхание",
            "49. Приступы жара или озноба",
            "50. Необходимость избегать некоторых мест или действий, т.к. они Вас пугают",
            "51. То, что Вы легко теряете мысль",
            "52. Онемение или покалывание в различных частях тела",
            "53. Комок в горле",
            "54. Ощущение, что будущее безнадежно",
            "55. То, что Вам трудно сосредоточиться",
            "56. Ощущение слабости в разных частях тела",
            "57. Ощущение напряженности или взвинченности",
            "58. Тяжесть в конечностях",
            "59. Мысли о смерти",
            "60. Переедание",
            "61. Ощущение неловкости, когда люди наблюдают за Вами или говорят о Вас",
            "62. То, что у Вас в голове чужие мысли",
            "63. Импульсы причинять телесные повреждения или вред кому-либо",
            "64. Бессонница по утрам",
            "65. Потребность повторять действия: прикасаться, мыться, пересчитывать и т.д.",
            "66. Беспокойный и тревожный сон",
            "67. Импульсы ломать или крушить что-нибудь",
            "68. Наличие у вас идей или верований, которые не разделяют другие",
            "69. Чрезмерная застенчивость при общении с другими",
            "70. Чувство неловкости в людных местах (магазинах и т.п.)",
            "71. Чувство, что все, что бы Вы не делали, требует больших усилий",
            "72. Приступы ужаса или паники",
            "73. Чувство неловкости, когда Вы едите и пьете на людях",
            "74. То, что Вы часто вступаете в спор",
            "75. Нервозность, когда Вы оставались одни",
            "76. То, что другие недооценивают Ваши достижения",
            "77. Чувство одиночества, даже когда Вы с другими людьми",
            "78. Такое сильное беспокойство, что Вы не могли усидеть на месте",
            "79. Ощущение собственной никчемности",
            "80. Ощущение, что с Вами произойдет что-то плохое",
            "81. То, что Вы кричите или швыряетесь вещами",
            "82. Боязнь, что Вы упадете в обморок на людях",
            "83. Ощущение, что люди злоупотребляют Вашим доверием",
            "84. Нервировавшие Вас сексуальные мысли",
            "85. Мысль, что Вы должны быть наказаны за Ваши грехи",
            "86. Кошмарные мысли или видения",
            "87. Мысль о том, что с Вашим телом что-то не в порядке",
            "88. То, что Вы не чувствуете близости ни к кому",
            "89. Чувство вины",
            "90. Мысли о том, что с Вашим рассудком творится что-то неладное",
        ]

        self.options = [
            "Совсем нет",
            "Немного",
            "Умеренно",
            "Сильно",
            "Очень сильно"
        ]

        self.responses = {}

        self.question_label = Label(master, text=self.questions[0])
        self.question_label.pack(anchor='w')
        self.answer_var = StringVar(master)
        self.answer_var.set(self.options[0])
        for option in self.options:
            answer_button = Radiobutton(master, text=option, variable=self.answer_var, value=option)
            answer_button.pack(anchor='w')
        self.next_button = Button(master, text="Следующий вопрос", command=self.next_question)
        self.next_button.pack(pady=10,anchor='w')
        self.save_button = Button(master, text="Результаты", command=self.save_responses)
        self.save_button.pack(anchor='w')
        self.current_question = 0
        self.filename = "SCL-90.xlsx"

    def next_question(self):
        response = self.answer_var.get()
        answer_value = self.options.index(response)
        self.responses[self.current_question] = answer_value
        self.current_question += 1
        if self.current_question < len(self.questions):
            self.question_label.config(text=self.questions[self.current_question])
        else:
            messagebox.showinfo("Тест окончен", "Вы ответили на все вопросы.")
    def save_responses(self):
        # Получение информации о пациенте
        fio = name_entry.get()
        age = age_entry.get()
        gender = gender_var.get()
        if foto:
            # Создание PDF-файла
            pdf = SimpleDocTemplate('result.pdf', pagesize=letter)
            story = []
            styles = getSampleStyleSheet()  # дефолтовые стили
            styles['Normal'].fontName = 'DejaVuSerif'
            styles['Heading1'].fontName = 'DejaVuSerif'
            pdfmetrics.registerFont(TTFont('DejaVuSerif', 'DejaVuSerif.ttf', 'UTF-8'))
            story.append(Paragraph('Результаты пациента', styles["Normal"]))

            # Добавление фотографии в PDF
            image1 = Image(foto, width=200, height=200)
            story.append(image1)


            # Добавление информации о пациенте в PDF
            story.append(Spacer(1, 12))
            story.append(Paragraph(f'ФИО: {fio}', styles["Normal"]))
            story.append(Paragraph(f'Возраст: {age} лет', styles["Normal"]))
            story.append(Paragraph(f'Пол: {gender}', styles["Normal"]))

            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active

            for i in range(len(self.questions)):
                response = self.responses.get(i)
                if response is not None:
                    cell = sheet.cell(row=2 + i, column=3 + response)
                    cell.value = response

            workbook.save(self.filename)
            workbook.close()
            excel_app = xlwings.App(visible=False)
            excel_book = excel_app.books.open('SCL-90.xlsx')
            excel_book.save()
            excel_book.close()
            excel_app.quit()
            workbook = openpyxl.load_workbook('SCL-90.xlsx', data_only=True)
            worksheet = workbook.active


            # Создание пустых списков для хранения данных
            names = []
            values = []
            colors = []

            # Получение значений из нужных столбцов и строк в обратном порядке
            for i in range(106, 94, -1):
                name = worksheet.cell(row=i, column=2).value
                value = worksheet.cell(row=i, column=3).value
                names.append(name)
                values.append(float(value))  # Преобразование значения в число

            # Задание цветов в зависимости от условий
            for i in range(len(names)):
                if names[i] == 'Общий индекс тяжести':
                    colors.append('red' if values[i] > 1.2 else 'green')
                elif names[i] == 'Общее число положительных ответов':
                    colors.append('red' if values[i] > 51 else 'green')
                elif names[i] == 'Индекс наличия симптоматического дистресса':
                    colors.append('red' if values[i] > 1.81 else 'green')
                elif names[i] == 'Шкала соматизации':
                    colors.append('red' if values[i] > 1.01 else 'green')
                elif names[i] == 'Шкала обсессивности-компульсивности':
                    colors.append('red' if values[i] > 1.31 else 'green')
                elif names[i] == 'Шкала межличностной сензитивности':
                    colors.append('red' if values[i] > 1.61 else 'green')
                elif names[i] == 'Шкала депрессии':
                    colors.append('red' if values[i] > 1.31 else 'green')
                elif names[i] == 'Шкала тревожности':
                    colors.append('red' if values[i] > 1.11 else 'green')
                elif names[i] == 'Шкала враждебности':
                    colors.append('red' if values[i] > 1.41 else 'green')
                elif names[i] == 'Шкала фобии':
                    colors.append('red' if values[i] > 0.71 else 'green')
                elif names[i] == 'Шкала паранойяльных тенденций':
                    colors.append('red' if values[i] > 1.31 else 'green')
                elif names[i] == 'Шкала психотизма':
                    colors.append('red' if values[i] > 0.91 else 'green')

            plt.figure(figsize=(8, 6))  # Размер окна графика (ширина, высота)
            bars = plt.barh(names, values, color=colors)


            for i, bar in enumerate(bars):
                value = values[i]
                plt.text(value, i, f'{value:.2f}', ha='left', va='center')

            plt.gca().xaxis.set_major_formatter(FormatStrFormatter('%.2f'))
            plt.subplots_adjust(left=0.48)
            plt.title('Результаты тестирования')

            # Сохранение графика в PNG
            plt.savefig('graph.png')

            # Добавление графика в PDF
            graph_image = Image('graph.png', width=400, height=300)
            story.append(graph_image)

            # Создание PDF
            pdf.build(story)
            print('PDF-файл сохранен успешно.')

        else:
            print('Не удалось сохранить PDF-файл. Пожалуйста, загрузите фото.')

def browse_file():
    global foto
    filename = filedialog.askopenfilename()
    # Отобразить выбранное изображение на форме
    if filename:
        image = PILImage.open(filename)
        image = image.resize((200, 200), PILImage.BILINEAR)  # Изменение размера изображения
        photo = ImageTk.PhotoImage(image)
        image_label.configure(image=photo)
        image_label.image = photo
        foto = filename


def start_test():
    test = Test(test_frame)

root = tk.Tk()
root.title("Тестирование пациента")
root.geometry('950x500')

# Фрейм для информации о пациенте
patient_info_frame = tk.Frame(root)
patient_info_frame.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

# Фрейм для ввода ФИО
name_frame = tk.Frame(patient_info_frame)
name_frame.pack(pady=10, anchor='w')
name_label = tk.Label(name_frame, text="ФИО пациента:")
name_label.pack(side=tk.LEFT)
name_entry = tk.Entry(name_frame, width=50)
name_entry.pack(side=tk.LEFT)

# Фрейм для ввода возраста
age_frame = tk.Frame(patient_info_frame)
age_frame.pack(pady=10, anchor='w')
age_label = tk.Label(age_frame, text="Возраст:")
age_label.pack(side=tk.LEFT)
age_entry = tk.Entry(age_frame)
age_entry.pack(side=tk.LEFT)

# Фрейм для ввода пола
gender_frame = tk.Frame(patient_info_frame)
gender_frame.pack(pady=10, anchor='w')
gender_label = tk.Label(gender_frame, text="Пол:")
gender_label.pack(side=tk.LEFT)
gender_var = tk.StringVar()
gender_var.set(0)
male_radio = tk.Radiobutton(gender_frame, text="Мужской", variable=gender_var, value="Мужской")
male_radio.pack(side=tk.LEFT)
female_radio = tk.Radiobutton(gender_frame, text="Женский", variable=gender_var, value="Женский")
female_radio.pack(side=tk.LEFT)
# Фрейм для загрузки фото
photo_frame = tk.Frame(patient_info_frame)
photo_frame.pack(pady=10, anchor='w')
photo_label = tk.Label(photo_frame, text="Фото: ")
photo_label.pack(side=tk.LEFT)
upload_button = tk.Button(photo_frame, text="Загрузить", command=browse_file)
upload_button.pack(side=tk.LEFT)
file_path_label = tk.Label(photo_frame, text="")
file_path_label.pack(side=tk.LEFT)

# Фрейм для кнопки тестирования
test_button_frame = tk.Frame(patient_info_frame)
test_button_frame.pack(pady=10, anchor='w')
test_button = tk.Button(test_button_frame, text="Тестирование", command=start_test)
test_button.pack()

# Фрейм для отображения загруженного изображения
image_frame = tk.Frame(patient_info_frame)
image_frame.pack(pady=10, anchor='w')
image_label = tk.Label(image_frame)
image_label.pack()

# Фрейм для тестирования
test_frame = tk.Frame(root)
test_frame.pack(side=tk.LEFT, padx=10, pady=10, fill=tk.BOTH, expand=True, anchor='ne')

root.mainloop()